# -*- coding: utf-8 -*-
"""AMPS 뉴스레터 대시보드 — 샘플/업로드 양식 엑셀 생성 스크립트.

실행:  python3 make_samples.py
생성물:
  templates/뉴스레터_업로드양식.xlsx   (헤더 + 작성안내 시트)
  templates/뉴스레터_샘플데이터.xlsx   (3~7월, HYUNDAI/DEVELON, 12개 지역 샘플)
  js/sample-data.js                    (동일 데이터를 브라우저에 임베드)

* 모든 이메일은 가짜(example.com) 주소입니다.
"""
import json
import os
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
random.seed(42)

REGIONS = ["본사", "중남미", "아프리카", "중동", "러시아/CIS", "아시아",
           "대양주", "터키/이스라엘", "CNHI", "브라질", "인도네시아", "북미"]
BRANDS = ["HYUNDAI", "DEVELON"]
MONTHS = [3, 4, 5, 6, 7]

# 지역별 딜러 기본 수신 규모 / 기본 오픈율 (스크린샷 수치와 유사한 스케일)
DEALER_BASE = {
    #            수신(H, D)   기본 오픈율
    "중남미":      (249, 180, 0.35),
    "아프리카":    (122, 95, 0.28),
    "중동":        (75, 60, 0.46),
    "러시아/CIS":  (56, 48, 0.19),
    "아시아":      (62, 70, 0.26),
    "대양주":      (28, 40, 0.20),
    "터키/이스라엘": (22, 18, 0.41),
    "CNHI":        (40, 55, 0.24),
    "브라질":      (5, 12, 0.22),
    "인도네시아":  (35, 52, 0.30),
    "북미":        (90, 85, 0.27),
}
# 내부직원은 본사에만 존재
INTERNAL_BASE = {"HYUNDAI": 180, "DEVELON": 120}

# 월별 브랜드 오픈율 추세 계수 (스크린샷 라인 형태: H 36→28→32 / D 39→29)
TREND = {
    "HYUNDAI": {3: 1.12, 4: 1.04, 5: 0.98, 6: 0.88, 7: 0.99},
    "DEVELON": {3: 1.25, 4: 1.22, 5: 1.01, 6: 0.93, 7: 0.91},
}

CONTENTS = {
    "HYUNDAI": ["신제품 HW160A 런칭", "Tilt Link 기능 소개", "글로벌 서비스 교육 캠페인",
                "Parts 프로모션 안내", "고객 성공 스토리", "전시회(bauma) 소식"],
    "DEVELON": ["DX 시리즈 업데이트", "Develon Days 리캡", "스마트 솔루션 소개",
                "부품 특가 프로모션", "딜러 트레이닝 아카데미", "지속가능경영 뉴스"],
}

READERS = (
    [f"dealer{i:02d}@example.com" for i in range(1, 13)]
    + [f"staff{i:02d}@example.com" for i in range(1, 5)]
    + [f"partner{i:02d}@example.com" for i in range(1, 5)]
)  # 총 20명 — 전부 가짜 주소


def jitter(v, pct):
    return v * (1 + random.uniform(-pct, pct))


def build_sends():
    rows = []
    for brand in BRANDS:
        for m in MONTHS:
            # 딜러 (11개 지역)
            for region, (h, d, rate) in DEALER_BASE.items():
                recv = max(3, round(jitter(h if brand == "HYUNDAI" else d, 0.10)))
                orate = min(0.62, max(0.08, jitter(rate * TREND[brand][m], 0.12)))
                opens = min(recv, round(recv * orate))
                clicks = round(opens * random.uniform(0.28, 0.50))
                rows.append([m, brand, "딜러", region, recv, opens, clicks])
            # 내부직원 (본사)
            recv = round(jitter(INTERNAL_BASE[brand], 0.06))
            orate = min(0.85, jitter(0.55 * TREND[brand][m], 0.08))
            opens = min(recv, round(recv * orate))
            clicks = round(opens * random.uniform(0.30, 0.45))
            rows.append([m, brand, "내부직원", "본사", recv, opens, clicks])
    return rows


def build_contents():
    rows = []
    for brand in BRANDS:
        for m in MONTHS:
            picks = random.sample(CONTENTS[brand], k=random.choice([4, 5, 6]))
            for c in picks:
                rows.append([m, brand, c, random.randint(3, 14)])
    return rows


def build_readers():
    rows = []
    for brand in BRANDS:
        for m in MONTHS:
            picks = random.sample(READERS, k=random.randint(9, 14))
            for email in picks:
                # 소수의 VIP 독자가 높은 클릭을 갖도록
                heavy = email in ("dealer01@example.com", "dealer05@example.com",
                                  "partner02@example.com")
                clicks = random.randint(6, 21) if heavy else random.randint(1, 7)
                rows.append([m, brand, email, clicks])
    return rows


SENDS_HEADER = ["월", "브랜드", "대상", "지역", "수신자수", "오픈수", "클릭수"]
CONTENT_HEADER = ["월", "브랜드", "콘텐츠명", "클릭수"]
READER_HEADER = ["월", "브랜드", "이메일", "클릭수"]

GUIDE_ROWS = [
    ["AMPS 뉴스레터 대시보드 — 엑셀 업로드 양식 작성 안내", "", ""],
    ["", "", ""],
    ["시트", "컬럼", "설명 / 허용 값"],
    ["발송실적", "월", "숫자(3) 또는 '3월' 형식. 예: 3, 4월, 2025-05 모두 인식"],
    ["발송실적", "브랜드", "HYUNDAI 또는 DEVELON (Hyundai/현대, Develon/디벨론 표기 허용)"],
    ["발송실적", "대상", "내부직원 또는 딜러"],
    ["발송실적", "지역", "본사, 중남미, 아프리카, 중동, 러시아/CIS, 아시아, 대양주, 터키/이스라엘, CNHI, 브라질, 인도네시아, 북미"],
    ["발송실적", "수신자수", "해당 월·지역·대상에게 발송(수신)된 인원 수 (숫자)"],
    ["발송실적", "오픈수", "메일을 오픈한 고유 인원 수 (숫자)"],
    ["발송실적", "클릭수", "본문 링크 클릭 수 (숫자)"],
    ["콘텐츠클릭", "월/브랜드/콘텐츠명/클릭수", "콘텐츠(기사)별 클릭 집계"],
    ["독자클릭", "월/브랜드/이메일/클릭수", "독자별 클릭 집계 (VIP 독자 분석용)"],
    ["", "", ""],
    ["※ 컬럼명은 유사 표기(예: '오픈 수', 'Opens', '수신자 수')도 자동 인식됩니다.", "", ""],
    ["※ 시트명도 유사 표기(예: 'Sends', '발송', 'Contents', 'Readers')를 인식합니다.", "", ""],
]


def style_header(ws, ncols):
    fill = PatternFill("solid", fgColor="1B3A6B")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = 14


def write_sheet(wb, name, header, rows):
    ws = wb.create_sheet(name)
    ws.append(header)
    for r in rows:
        ws.append(r)
    style_header(ws, len(header))
    return ws


def make_workbook(path, sends, contents, readers, with_guide):
    wb = Workbook()
    wb.remove(wb.active)
    if with_guide:
        ws = wb.create_sheet("작성안내")
        for r in GUIDE_ROWS:
            ws.append(r)
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 90
        ws["A1"].font = Font(bold=True, size=13)
        ws["A3"].font = ws["B3"].font = ws["C3"].font = Font(bold=True)
    write_sheet(wb, "발송실적", SENDS_HEADER, sends)
    write_sheet(wb, "콘텐츠클릭", CONTENT_HEADER, contents)
    write_sheet(wb, "독자클릭", READER_HEADER, readers)
    wb.save(path)
    print("saved:", path)


def main():
    os.makedirs(os.path.join(BASE, "templates"), exist_ok=True)
    sends = build_sends()
    contents = build_contents()
    readers = build_readers()

    # 1) 업로드 양식 (헤더 + 예시 2행 + 작성안내)
    make_workbook(
        os.path.join(BASE, "templates", "뉴스레터_업로드양식.xlsx"),
        sends[:2], contents[:2], readers[:2], with_guide=True)

    # 2) 샘플 데이터
    make_workbook(
        os.path.join(BASE, "templates", "뉴스레터_샘플데이터.xlsx"),
        sends, contents, readers, with_guide=False)

    # 3) js/sample-data.js 임베드 (동일 데이터)
    data = {
        "sends": [dict(zip(["month", "brand", "audience", "region",
                            "recipients", "opens", "clicks"], r)) for r in sends],
        "contents": [dict(zip(["month", "brand", "content", "clicks"], r))
                     for r in contents],
        "readers": [dict(zip(["month", "brand", "email", "clicks"], r))
                    for r in readers],
    }
    js = ("// 자동 생성 파일 — make_samples.py 실행으로 재생성됩니다. 직접 수정하지 마세요.\n"
          "// 모든 이메일은 가짜(example.com) 주소입니다.\n"
          "var SAMPLE_DATA = "
          + json.dumps(data, ensure_ascii=False, indent=1)
          + ";\n\nif (typeof module !== 'undefined' && module.exports) { module.exports = SAMPLE_DATA; }\n"
          "if (typeof window !== 'undefined') { window.SAMPLE_DATA = SAMPLE_DATA; }\n")
    out = os.path.join(BASE, "js", "sample-data.js")
    with open(out, "w", encoding="utf-8") as f:
        f.write(js)
    print("saved:", out, f"(발송 {len(sends)}행 / 콘텐츠 {len(contents)}행 / 독자 {len(readers)}행)")


if __name__ == "__main__":
    main()
